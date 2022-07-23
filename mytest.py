import pandas as pd
from openpyxl import load_workbook
import openpyxl

doc = 'D:/projects/Email_Book.xlsx'
# wb = openpyxl.load_workbook(doc)
# ws = wb.active
# print("Maximum rows before removing:", ws.max_row)
# ws.delete_rows(2, ws.max_row-1)
# print("Maximum rows after removing:", ws.max_row)

df = pd.DataFrame({'Emails ' : ['arun123@gmail.com','arun234@gmail.com','babu3646@gmail.com']})
writer = pd.ExcelWriter(doc, engine='openpyxl')
# try to open an existing workbook
writer.book = load_workbook(doc)
# copy existing sheets
writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
reader = pd.read_excel(r'doc')
# write out the new sheet
df.to_excel(writer,index=False,header=False,startrow=len(reader)+1)

writer.close()


# import openpyxl library
import openpyxl

# function to remove empty rows

def remove(sheet, row):
	
	for cell in row:
		# check the value of each cell in
		# the row, if any of the value is not
		# None return without removing the row
		if cell.value != None:
			return
	# get the row number from the first cell
	# and remove the row
	sheet.delete_rows(row[0].row, 1)
if __name__ == '__main__':
	# enter your file path
	path = doc
	# load excel file
	wb = openpyxl.load_workbook(path)
	# select the sheet
	ws = wb['Sheet1']
	print("Maximum rows before removing:", ws.max_row)
	# iterate the sheet object
	for row in ws:
	remove(ws,row)
	print("Maximum rows after removing:",sheet.max_row)
	# save the file to the path
	path = './openpy.xlsx'
	book.save(path)
